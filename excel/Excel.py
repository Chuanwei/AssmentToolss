#encoding:utf-8
"""
@author:Liod
@file:Excel.py.py
@time:17-12-24下午9:15
"""
# encoding:utf-8
"""
@author:Liod
@file:Excel.py
@time:17-12-22上午11:09
"""
import platform
import openpyxl
from openpyxl import load_workbook
import Libs
import os
import Color
import time
from Libs.OS.Linux import *
from Color.OS_Check import OS


class ExceWrite(object):
    def __init__(self, modile_name='', save_name=''):
        self.module = modile_name
        self.savename = save_name
        self.main = load_workbook(modile_name)
        self.sheet = self.main.get_sheet_by_name(self.main.sheetnames[0])
        self.max_row = self.sheet.max_row


    def GetIndex(self, B=''):
        """
        :param B: 
        :return: index [] 
        """
        result = []
        for i in range(1, self.max_row):
            if self.sheet["%s%s" % (B, str(i))].value != None:
                result.append(i)
        return result

    def GetTuple(self, B=''):
        """

        :param B: 
        :return:[(), ()] 
        """
        result = []
        index = self.GetIndex(B)
        for p in range(0, len(index)):
            try:
                if index[p + 1] - index[p] == 0:
                    tu = (index[p], index[p])
                    result.append(tu)
                else:
                    tu = (index[p], index[p + 1] - 1)
                    result.append(tu)
            except IndexError:
                tu = (index[p], index[p])
                result.append(tu)
        return result

    def Create_dict(self):
        result = {}
        A = self.GetTuple('B')
        D = self.GetTuple('D')
        for re in D:
            for e in A:
                if e[1] == re[-1]:
                    try:
                        result["%s" % self.sheet["D%s" % e[0]].value] = D[:D.index(re) + 1]
                    except IndexError:
                        result["%s" % self.sheet["D%s" % e[0]].value] = D[D.index(re):]
        for i in result.keys():
            print i

    def SaveTuple(self, dict={}):
        """

        :param dict:
         根据dict索引写入内容，更改内容，主要设计测评表Ｄ列根据前面的算法索引跌代更改数据
        :return: 
        """
        test = {}
        tup = self.GetTuple("B")
        for t in tup:
            test["%s" % self.sheet["C%s" % t[0]].value] = ""
            print test
        for t in test.keys():
            print t


class Excel_Create(object):

    def __init__(self, loadtag, savename=""):
        self.wb = load_workbook("{0}{1}.xlsx".format(self.getOsDir()[0], loadtag))
        self.sheet = self.wb.worksheets[0]
        self.savename = "{0}{1}.xlsx".format(self.getOsDir()[1], savename)
        self.col = OS().get_col()


    def getOsDir(self):
        if 'Windows' in platform.system():
            getSeparator = '\\'
        else:
            getSeparator = '/'
        modiles = os.getcwd() + getSeparator + 'modiles' + getSeparator + "DATABASE" + getSeparator
        day = time.strftime('%Y-%m-%d', time.localtime(time.time()))
        osdir = os.getcwd() + getSeparator + day + getSeparator
        if os.path.exists(osdir):
            result = os.getcwd() + getSeparator + day +  getSeparator
        else:
            try:
                os.mkdir(osdir)
                result =  os.getcwd() + getSeparator + day +  getSeparator
            except:
                print u"文件夹创建失败请检查权限"
                result = modiles
                print result
        return (modiles, result)

    def result_introduction(self, row=[2, 0], celdata={}):

        row[1] = len(celdata) + 1
        for i in range(row[0], row[1] + 1):
            self.sheet["D%s" % i].value = celdata["D%s" % i]["content"]
            self.sheet["C%s" % i].value = celdata["D%s" % i]["command"]
        print self.savename
        self.wb.save(self.savename)
        # try:
        #     self.wb.save(self.savename)
        #     self.col.printBlue(u"生成文件成功！文件路径为： %s \n"%self.savename)
        # except Exception as e:
        #     self.col.printRed(u"生成文件失败， 错误原因： %s \n"%e)


aaa = Excel_Create("mysql", "aa")
print aaa.savename
